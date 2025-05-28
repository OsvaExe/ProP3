# Archivo de excel Empleados.xlsx
import openpyxl
import os

# Workbook o libro de excel:
directorio = os.path.dirname(__file__)  # con este comando se consigue la direccion de main
excel_path = os.path.join(directorio, "Empleados.xlsx")
wb = openpyxl.load_workbook(excel_path)
# WorkSheet u Hoja activa de Excel
ws = wb.active

def mostrar_info_creadores():
    print("\n=== INFORMACIÓN DEL PROGRAMA ===")
    print("Nombre del equipo: Los mocosos")
    print("Desarrollado por:")
    print("- Armando Velazquez")
    print("- Emilio Gallardo")
    print("- Osvaldo Hernandez")
    print("- Luis Javier Torres")
    input("\nPresiona Enter para volver al menú...")

def validar():
    """
    Valida que el usuario ingrese un número entero entre 1 y 7.
    Continúa pidiendo el input hasta que se ingrese un valor válido.
    """
    while True:
        valor = input("Ingresa un número entre 1 a 7 \n:")
        try:
            numero = int(valor)
            if 1 <= numero <= 7:
                return numero
            else:
                print("El número debe estar entre 1 y 7")
        except ValueError:
            print("¡Debe ser un número entero!")

def ppal(): 
    opcion = 0
    while opcion != 7:  
        print("\n=== MENÚ PRINCIPAL ===")
        print("1. Alta")
        print("2. Modificación")
        print("3. Dar de baja un empleado")
        print("4. Consulta de empleado")
        print("5. Sanciones y actualización de sueldo por quincena")
        print("6. Información del programa")  
        print("7. Salir del programa")
        
        opcion = validar()

        match(opcion):
            #--------------------AGREGAR EMPLEADO----------------------    
            case(1):
                conteof = 0
                numrep = True  # bandera para confirmar nombre repetido
                empleadoData = []  # la pase aqui para asegurar que la lista este vacia
                # Verificación de la cantidad de datos
                CantMAXfilas = ws.max_row
                CantMAXcol = ws.max_column
                print("Filas: ", CantMAXfilas, " Columnas: ", CantMAXcol)
                
                print("Se agregará un nuevo empleado al registro")
                numeroe = int(input("Escriba el numero del empleado:"))
                # confirmar que el numero no este en la base de datos
                col = ws['A']
                for cell in col:
                    if str(cell.value) == str(numeroe):
                        print("Este número de empleado ya existe en la base de datos")
                        numrep = False
                if numrep:
                    nombres = input("Escriba el nombre o nombres (no apellidos) del empleado:")
                    apellidos = input("Escriba el o los apellidos del empleado: ")
                    print("Se agregó a: ", numeroe, nombres, apellidos)
                    empleadoData.append(numeroe)  # NUMERO EMPLEADO
                    empleadoData.append(nombres)  # NOMBRE
                    empleadoData.append(apellidos)  # APELLIDO
                    empleadoData.append("A")  # ESTATUS
                    empleadoData.append(1)  # ASISTENCIA 
                    empleadoData.append(0)  # RETARDO
                    print(empleadoData)
                    ws.append(empleadoData)
                    wb.save(excel_path)  # se guardan los cambios en el archivo
            
            #----------MODIFICAR ELEMENTO--------------------------------------------------------------------        
            case(2):
                #Modificar un elemento
                CantMAXfilas = ws.max_row
                CantMAXcol = ws.max_column
                conteof = 0
                print("Modificarás un elemento")
                print("Recordar que el número de empleado no se podrá modificar")
                numeroe = input("Numero de empleado: ")
                for row in ws:
                    num_emplea = row[0].value #primer elemento de una fila
                    conteof = conteof + 1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                    if num_emplea == numeroe:
                        print("El empleado se encontró")
                        numfila = conteof
                        print("¿Desea modificar...")
                        print("1. Nombre(s)")
                        print("2. Apellidos")
                        print("3. Corregir retraso")
                        print("4. Cambiar Estatus")
                        OPmenumodif = int(input("Seleccione una opción a modificar: "))
                        match(OPmenumodif):
                            case(1): 
                                #Modificar nombre
                                nombrenuevo = input("Teclee el(los) nuevo(s) nombre(s): ")
                                ws.cell(row=numfila, column=2, value=nombrenuevo)
                                wb.save(excel_path)  # Se guardan los cambios en el archivo
                                print("El nombre del empleado ", numeroe, " se actualizó a ")
                                valornuevo = ws.cell(row=numfila, column=2).value
                                print(valornuevo)
                                print("\U0001F44D")
                                break
                            case(2):
                                #Modificar Apellidos
                                apellidonuevo = input("Teclee el(los) nuevo(s) apellido(s): ")
                                ws.cell(row=numfila, column=3, value=apellidonuevo)
                                wb.save(excel_path)  # Se guardan los cambios en el archivo
                                print("El(los) apellido(s) del empleado ", numeroe, " se actualizó a: ")
                                valornuevo = ws.cell(row=numfila, column=3).value
                                print(valornuevo)
                                print("\U0001F44D")
                                break   
                            case(3):
                                #Modificar retardos
                                print("Modificar la cantidad de retrasos, esto sucede cuando se justifican")
                                print("La cantidad de retrasos es de 1 a 3")
                                print("\U0001F480") 
                                break
                            case(4):
                                #Modificar Estatus
                                estatus_actual = ws.cell(row=numfila, column=4).value
                                nuevo_estatus = input("Cambiar estatus (solo se permite 'A' o 'I'): ").upper()

                                if nuevo_estatus not in ["A", "I"]:
                                    print("Error: El estatus solo puede ser 'A' o 'I'")
                                    break

                                if nuevo_estatus == estatus_actual:
                                    print(f"No se hará nada. El estatus ya es '{estatus_actual}'")
                                    break

                                ws.cell(row=numfila, column=4, value=nuevo_estatus)
                                wb.save(excel_path)
                                print(f"Estatus del empleado {numeroe} actualizado a: {nuevo_estatus}")
                                break

                            case _:
                                print("Opción no válida")
                                break
                if conteof == CantMAXfilas:
                    print("Empleado no encontrado \U0001F633")
            
            #-----------BAJA-----------------------------------------------------------------
            case(3): 
                #BAJA de empleado
                conteof = 0
                print("Eliminarás un empleado")
                numeroe = input("Numero de empleado a eliminar \U0001F631: ")
                for row in ws:
                    num_emplea = row[0].value #primer elemento de una fila
                    conteof = conteof + 1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                    if num_emplea == numeroe:
                        print("El empleado se encontró")
                        numfila = conteof
                        ws.cell(row=numfila, column=4, value="B") #"B=BAJA" se asigna un valor en la columna 4
                        wb.save(excel_path)  # Se guardan los cambios en el archivo
                        valornuevo = ws.cell(row=numfila, column=4).value #Se obtiene el valor que se guardo en ESTATUS
                        print("El estatus del empleado ", numeroe, " se actualizó a: ", valornuevo)
                        print("\U0001F494")
                        break
                else:
                    print("Empleado no encontrado \U0001F633")
            
            #--------------------CONSULTA-----------------------------------------------------------        
            case(4): #Consulta de empleados
                print("\U0001FAF8:")
                CantMAXfilas = ws.max_row
                CantMAXcol = ws.max_column
                conteof = 0
                print("Consulta")
                numeroe = input("Numero de empleado a consultar \U0001F9D0: ")
                for row in ws:
                    num_emplea = row[0].value #primer elemento de una fila
                    conteof = conteof + 1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                    if num_emplea == numeroe:
                        print("El empleado se encontró")
                        numfila = conteof
                        break #Ya encontrado el empleado se interrumpe el for
                else:
                    print("Empleado no encontrado \U0001F633")
                    continue
                
                for i in range(1, CantMAXcol + 1): #Desde 1 hasta max cantidad de columnas (+1 para inclusivo)
                    valores_empleado = ws.cell(row=numfila, column=i).value 
                    print(valores_empleado)
            
            #--------------SANCIONES---------------------------------------------------------------------------
            case(5): #Sanciones
                #Se buscarán aquellos que tengan de 3 retardos y se quitará 10% de su sueldo
                print("\U0001FAE2")
                CantMAXfilas = ws.max_row
                CantMAXcol = ws.max_column
                print("Actualizar sanciones")
                for i in range(2, CantMAXfilas + 1): #Desde 2 hasta max filas (+1 para inclusivo)
                    cant_retardos = ws.cell(row=i, column=6).value #columna de retardos
                    if cant_retardos == 3:
                        num_emplea = ws.cell(row=i, column=1).value
                        nombre_emplea = ws.cell(row=i, column=2).value
                        apellidos_emplea = ws.cell(row=i, column=3).value
                        ws.cell(row=i, column=7, value=0.1) #descuento
                        sueldo_actual = 21000 - (21000 * 0.1)
                        ws.cell(row=i, column=8, value=sueldo_actual)
                        wb.save(excel_path)  # Se guardan los cambios en el archivo
                        sueldo_nuevo = ws.cell(row=i, column=8).value 
                        print("El empleado ", num_emplea, " ", nombre_emplea, " ", apellidos_emplea, " tiene de sueldo: ", sueldo_nuevo)
                    else:
                        print(f"Empleado {ws.cell(row=i, column=1).value} no tiene sanción (retardos: {cant_retardos})")
            
            case(6):  
                mostrar_info_creadores()
            
            case(7):
                print("¡Saliendo del programa!")
                break
            
            case _:
                print("Opción no válida. Intente de nuevo.")

if __name__ == "__main__":
    ppal()
